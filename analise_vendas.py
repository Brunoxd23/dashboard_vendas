import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Lendo o arquivo CSV
df = pd.read_csv('arquivo_vendas.csv')

# Calculando o faturamento
df['faturamento'] = df['quantidade'] * df['preco_unitario']

# Agrupando por produto e somando o faturamento
faturamento_por_produto = df.groupby('produto')['faturamento'].sum().reset_index()

# Encontrando o produto com maior e menor faturamento
produto_maior_faturamento = faturamento_por_produto.loc[faturamento_por_produto['faturamento'].idxmax()]
produto_menor_faturamento = faturamento_por_produto.loc[faturamento_por_produto['faturamento'].idxmin()]

# Criando um DataFrame mais organizado para o relatório
relatorio = pd.DataFrame([
    ['Resumo de Vendas', ''],
    ['', ''],
    ['Faturamento por Produto:', ''],
    *[[f'- {row.produto}', f'R$ {row.faturamento:,.2f}'] for _, row in faturamento_por_produto.iterrows()],
    ['', ''],
    ['Produto com Maior Faturamento:', produto_maior_faturamento['produto']],
    ['Valor:', f'R$ {produto_maior_faturamento["faturamento"]:,.2f}'],
    ['', ''],
    ['Produto com Menor Faturamento:', produto_menor_faturamento['produto']],
    ['Valor:', f'R$ {produto_menor_faturamento["faturamento"]:,.2f}']
], columns=['Análise', 'Resultado'])

# Criando o Excel com formatação
with pd.ExcelWriter('relatorio_vendas.xlsx', engine='openpyxl') as writer:
    # Aba Resumo
    relatorio.to_excel(writer, sheet_name='Resumo', index=False)
    
    # Pegando a planilha para formatação
    ws = writer.sheets['Resumo']
    
    # Formatação geral
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 35
    
    # Formatando o título
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Formatando cabeçalhos das seções
    for row in [3, 6, 9]:  # Linhas com cabeçalhos
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # Aba com faturamento por produto
    faturamento_formatado = faturamento_por_produto.sort_values('faturamento', ascending=False)
    faturamento_formatado['faturamento'] = faturamento_formatado['faturamento'].apply(lambda x: f'R$ {x:,.2f}')
    faturamento_formatado.columns = ['Produto', 'Faturamento']
    
    faturamento_formatado.to_excel(
        writer, 
        sheet_name='Faturamento por Produto',
        index=False
    )
    
    # Formatando a aba de faturamento
    ws = writer.sheets['Faturamento por Produto']
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 35
    
    # Formatando cabeçalhos
    for col in ['A1', 'B1']:
        ws[col].font = Font(bold=True)
        ws[col].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # Aba com dados brutos
    df.to_excel(writer, sheet_name='Dados Brutos', index=False)
    
    # Formatando a aba de dados brutos
    ws = writer.sheets['Dados Brutos']
    for col in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
        ws[f'{get_column_letter(col)}1'].font = Font(bold=True)
        ws[f'{get_column_letter(col)}1'].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

print("Arquivo 'relatorio_vendas.xlsx' foi criado com sucesso!")

# Exibindo os resultados no console
print("\nFaturamento por produto:")
print(faturamento_por_produto)
print(f"\nProduto com maior faturamento: {produto_maior_faturamento['produto']}")
print(f"Valor do maior faturamento: R$ {produto_maior_faturamento['faturamento']:,.2f}")
print(f"\nProduto com menor faturamento: {produto_menor_faturamento['produto']}")
print(f"Valor do menor faturamento: R$ {produto_menor_faturamento['faturamento']:,.2f}")